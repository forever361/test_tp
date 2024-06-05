# coding=utf-8
import os
import sys

from app.data2_check.commom.Constant_t import Constant_id

configPath = os.path.abspath(os.path.join(os.path.dirname(__file__),"../.."))
sys.path.append(configPath)
from app.useDB import ConnectSQL

userid = Constant_id().cookie_id
base_path = os.path.abspath(os.path.join(os.path.dirname(__file__),"../"))
basepath = os.path.abspath(os.path.join(os.path.dirname(__file__), "../.."))
userPath = os.path.join(basepath + "/userinfo/{}".format(userid))

# print(userPath)

import openpyxl
import xlrd
from xlutils.copy import copy
import time
import pandas as pd
from app.application import app


class ExcelUtilAll:
    def __init__(self):
        self.row = 1
        self.row_check = 1
        self.count_row = 1
        self.count_success = 0
        self.count_failure =0
        self.count_total = 0
        self.value_success = 0
        self.value_failure = 0
        self.value_total = 0

    def head_row(self):
        self.write_value_report(0, 1, 1, 'Source Table')
        self.write_value_report(0, 1, 2, 'Target Table')
        self.write_value_report(0, 1, 3, 'Source Count')
        self.write_value_report(0, 1, 4, 'Target Count')
        self.write_value_report(0, 1, 5, 'Count Result')
        self.write_value_report(0, 1, 6, 'Value Result')
        self.write_value_report(0, 1, 7, 'Value detail')

    def load_excel(self):
        '''Load excel '''
        user_id = Constant_id().cookie_id
        basepath = os.path.abspath(os.path.join(os.path.dirname(__file__), "../.."))
        # print(111111,basepath + "/userinfo/{}".format(userid))
        userPath = os.path.join(basepath + "/userinfo/{}".format(userid))
        folder_path = os.path.join(app.root_path, 'static', 'user_files', user_id)
        try:
            open_excel = openpyxl.load_workbook(folder_path + '/config/verification_result.xlsx')
            return open_excel

        except Exception as e:
            print("访问 sheetnames 时出错：", e)

    def get_data(self,index=None):
        '''Load the all content '''
        sheet_name = self.load_excel().sheetnames
        if index == None or index == "0":
            index = 0
        elif index >= 1:
            index = index
        data = self.load_excel()[sheet_name[index]]
        return data

    def get_sheet_num(self):
        ''' get the sheet number '''
        sheet_name = self.load_excel().sheetnames
        sheet_num = len(sheet_name)
        return

    def get_lines(self,key):
        ''' Get the all rows '''
        row = self.get_data(key).max_row
        return row

    def get_rows_value(self,key,row):
        ''' Get the rows value '''
        row_list = []
        for i in self.get_data(key)[row]:
            row_list.append(i.value)
        return row_list

    def get_col_value(self,key,row,col):
        ''' get the contents of a cell '''
        data = self.get_data(key).cell(row=row,column=col).value
        return data

    def write_value_report(self,key,row,col,value):
        ''' write data in the workbook '''
        user_id = Constant_id().cookie_id
        basepath = os.path.abspath(os.path.join(os.path.dirname(__file__), "../.."))
        userPath = os.path.join(basepath + "/userinfo/{}".format(userid))
        folder_path = os.path.join(app.root_path, 'static', 'user_files', user_id)
        wb = self.load_excel()
        wa = wb.sheetnames[key]
        wr = wb[wa]
        wr.cell(row,col,value)
        wb.save(folder_path + '/config/verification_result.xlsx')

    def get_diference_value(self,result):
        ''' result: diference value and times write excel '''
        self.row += 1
        if result:
            values = str(result)
            excel_data_all.write_value_report(0,self.row,3,values)
            excel_data_all.write_value_report(0,2,4,self.row-1)

    def get_check_count_times(self,count_check_flag=None):
        if count_check_flag == True:
            self.count_success += 1
            excel_data_all.write_value_report(0,2,7,self.count_success)
        elif count_check_flag == False:
            self.count_failure += 1
            excel_data_all.write_value_report(0,2,8,self.count_failure)


    def get_batch_count(self,s_tablename,t_tablename,s,t,row):
        row += 1
        if s_tablename and t_tablename:
            s_tablename = s_tablename.strip()
            t_tablename = t_tablename.strip()
            excel_data_all.write_value_report(0,row, 1, s_tablename)
            excel_data_all.write_value_report(0, row, 2, t_tablename)

        excel_data_all.write_value_report(0, row, 5, s)
        excel_data_all.write_value_report(0, row, 6, t)

    def get_value_result_pass(self,value_check_flag=None):
        if value_check_flag:
            self.value_success += 1
            excel_data_all.write_value_report(0, 2, 9, self.value_success)

#START
    def get_value_result_pass2(self,value_check_flag=None,times=2):
        if value_check_flag:
            self.value_success = 'Pass'
            excel_data_all.write_value_report(0, times, 6, self.value_success)
        else:
            self.value_success = 'Fail'
            excel_data_all.write_value_report(0, times, 6, self.value_success)

    def get_count_result_pass2(self,value_check_flag=None,times=2):
        if value_check_flag:
            self.value_success = 'Pass'
            excel_data_all.write_value_report(0, times, 5, self.value_success)
        else:
            self.value_success = 'Fail'
            excel_data_all.write_value_report(0, times, 5, self.value_success)

    def get_source_count(self,count,times):
        excel_data_all.write_value_report(0, times, 3, count)

    def get_target_count(self,count,times):
        excel_data_all.write_value_report(0, times, 4, count)

    def get_source_table_name(self,name,times):
        excel_data_all.write_value_report(0, times, 1, name)

    def get_target_table_name(self,name,times):
        excel_data_all.write_value_report(0, times, 2, name)

    def get_value_detail(self,name,times):
        excel_data_all.write_value_report(0, times, 7, name)

# END

    def get_value_result_fail(self):
        self.value_failure += 1
        excel_data_all.write_value_report(0, 2, 10, self.value_success)

    def init_excel_data(self,row):
        row += 1
        excel_data_all.write_value_report(0, row + 1, 1, '')
        excel_data_all.write_value_report(0, row + 1, 2, '')
        excel_data_all.write_value_report(0, row + 1, 3, '')
        excel_data_all.write_value_report(0, row + 1, 4, 0)
        excel_data_all.write_value_report(0, row + 1, 5, 0)
        excel_data_all.write_value_report(0, row + 1, 6, 0)
        excel_data_all.write_value_report(0, row + 1, 7, 0)
        excel_data_all.write_value_report(0, row + 1, 8, 0)
        excel_data_all.write_value_report(0, row + 1, 9, 0)
        excel_data_all.write_value_report(0, row + 1, 10, 0)
        excel_data_all.write_value_report(0, row + 1, 11, 0)


excel_data_all = ExcelUtilAll()


if __name__ == '__main__':
    handle = ExcelUtilAll()
    # handle.Data_html()
    # handle.get_check_count_pass(False)

